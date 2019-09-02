#!/usr/bin/env python3

'''
This script takes raw output from qPCR and performs calculations on the results.

Example:
    $ qPCR_calc.py file_name.xlsx CONTROL_NAME

Args:
    file_name: raw output from qPCR in xlsx format
    control_target: control unit to perform calculations against
        default = "GAPDH"

todo:
    - Get flow data
    - calculate 2^( (Flow_sample - Flow_gapdh) - (Sample_sample - Sample_gapdh) )
'''

import csv
import openpyxl
import sys
import pyexcel
import argparse


class SampleClass:
    def __init__(self, CT):
        self.CTtotal = float(CT)
        self.count = 1.0
    
    def add(self, CT):
        self.CTtotal = self.CTtotal + float(CT)
        self.count = self.count + 1.0

    def getAverage(self):
        return self.CTtotal / self.count

    def getCount(self):
        return self.count


def openExcel(excel_workbook, control_target, control_sample_list):
    data_dict = {} 
    wb_obj = openpyxl.load_workbook(excel_workbook)
    sheet_obj = wb_obj['Results']
    control_target_exist = False

    # iterate through spreadsheet rows til we find data
    start_process = False


    for row_obj_tup in sheet_obj.iter_rows(min_row = 40,  min_col=4, max_col=15):
        target = row_obj_tup[1].value
        sample = row_obj_tup[0].value
        ct = row_obj_tup[11].value

        # skip until we find Sample Name
        if (not start_process) and sample != 'Sample Name':
            continue
        elif sample == 'Sample Name':
            start_process = True
            print('The first row of data begins on ' + str(row_obj_tup[0].row+1))
            continue

        # stop if we reach line with no values
        if (start_process) and (target is None):
            print('The last row data ends on ' + str(row_obj_tup[0].row - 1) +'.')
            break

        # Check if control exist. control_target case insensitive
        if (not control_target_exist) and (target.upper() == control_target):
            control_target = target
            control_target_exist = True

        try:
            float(ct)
        except:
            print('Row ' + str(row_obj_tup[0].row) + ' does not have a numeric CT value. Row ignored.')
            continue

        #check if target exists in dict, then add sample to dict. else create dict
        if target in data_dict:
            # check if sample exists in subdict
            if sample in data_dict[target]:
                data_dict[target][sample].add(ct)
            else:
                data_dict[target][sample] = SampleClass(ct)

        else:
            # initialize key with subdict of sample and SampleClass object
            data_dict[target] = {sample: SampleClass(ct)}

    if control_target_exist == False:
        print('No control found in spreadsheet. No output created.')
        return

    control_dict = data_dict[control_target].copy()
    writeExcel(excel_workbook, data_dict, control_dict, control_target, control_sample_list)

def writeExcel(file_name, data_dict, control_dict, control_target, control_sample_list):
    wb_obj = openpyxl.Workbook()
    sheet_obj = wb_obj.active
    sheet_obj.title = 'Calculations'

    # if -tc used, initalize stuff
    if control_sample_list:
        #convert list to lower for comparison
        control_list = [x.lower() for x in control_sample_list]
        control_list.sort()
        list_dict = {}
        tracking_list = []
        print("Here is the list of control samples you've imported:", control_list)


    sheet_obj.cell(row = 1, column = 1).value = 'Target Name'
    sheet_obj.cell(row = 1, column = 2).value = 'Sample Name'
    sheet_obj.cell(row = 1, column = 3).value = 'Average CT'
    sheet_obj.cell(row = 1, column = 4).value = 'Delta CT'

    # start writing data on row 2
    row_index = 2

    # calculates control average and control delta
    for target_key in data_dict:
        for sample_key in data_dict[target_key]:
            ct_average = data_dict[target_key][sample_key].getAverage()
            ct_delta = ct_average - control_dict[sample_key].getAverage()

            # write
            sheet_obj.cell(row = row_index, column = 1).value = target_key
            sheet_obj.cell(row = row_index, column = 2).value = sample_key
            sheet_obj.cell(row = row_index, column = 3).value = ct_average
            sheet_obj.cell(row = row_index, column = 4).value = ct_delta

            row_index += 1

            # if -tc option used, start tracking delta ct
            if control_sample_list:
                sample_lower = sample_key.lower()
                if sample_lower in control_list:
                    if target_key in list_dict:
                        list_dict[target_key].add(ct_delta)
                    else:
                        list_dict[target_key] = SampleClass(ct_delta)
                    if sample_lower not in tracking_list: # for tracking purposes
                        tracking_list.append(sample_lower)
    
    # if -tc option used (if list exist), calculte target average, delta, and fold change
    if control_sample_list:

        sheet_obj.cell(row = 1, column = 5).value = 'Avg Dct CTRL'
        sheet_obj.cell(row = 1, column = 6).value = 'Normalize (Delta CT-Avg Dct CTRL)'
        sheet_obj.cell(row = 1, column = 7).value = 'Fold change (2^-Normalize)'

        row_index = 2
        for target_key in data_dict:
            for sample_key in data_dict[target_key]:
                dct_average = list_dict[target_key].getAverage()
                normalized = sheet_obj.cell(row = row_index, column = 4).value - dct_average
                fold_change = 2 ** -(normalized)

                sheet_obj.cell(row = row_index, column = 5).value = dct_average
                sheet_obj.cell(row = row_index, column = 6).value = normalized
                sheet_obj.cell(row = row_index, column = 7).value = fold_change

                row_index += 1

        tracking_list.sort()
        print("Here is the list of control samples we've found:", tracking_list)
        print("Here is the list of control samples not found:", [x for x in control_list if x not in tracking_list])

    new_file = 'CALCULATED_' + file_name
    wb_obj.save(new_file)
    print('Created file ' + new_file)

def main():
    parser = argparse.ArgumentParser(description="Performs calculations on qPCR results")
    parser.add_argument("-i", dest="input_file", help="qPCR input file in xls format, REQUIRED", required=True)
    parser.add_argument("-c", dest="control_target", help="target control name, default = GAPDH, OPTIONAL", default="GAPDH")
    parser.add_argument("-tc", dest="control_sample", help="list of sample controls, OPTIONAL", nargs='+')

    args = parser.parse_args()
    file_location = args.input_file
    control_target = args.control_target
    control_sample_list = args.control_sample
    
    try:
        # convert xls to xlsx
        if file_location.endswith('.xls'):
            temp_file = file_location + 'x'
            pyexcel.save_book_as(file_name = file_location, dest_file_name = temp_file)
            filename = temp_file
        else:
            filename = file_location

        openExcel(filename, control_target, control_sample_list)
    except openpyxl.utils.exceptions.InvalidFileException:
        print('Only xlsx files are supported.')
    

if __name__ == '__main__':
    main()